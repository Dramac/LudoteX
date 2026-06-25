"""
Exports du planning bénévole en Excel (.xlsx) et PDF.

Les deux exports partent de la grille produite par
`app.planning.services.construire_grille` (postes, jours, créneaux, besoins,
affectations) et la rendent sous une forme proche du tableur du bureau :

- Excel : UNE FEUILLE PAR JOUR, lignes = créneaux, colonnes = postes ; chaque
  case contient les noms affectés (cases sans besoin laissées vides/grisées).
- PDF : version imprimable/affichable (A4 paysage, une section par jour).

Dépendances : openpyxl (Excel) et reportlab (PDF), déjà au requirements.txt.
"""

from __future__ import annotations

from io import BytesIO

from app.services import format_local


def _heure(iso_utc: str | None) -> str:
    """Affiche un créneau en heure locale courte 'HH:MM' (depuis l'ISO UTC)."""
    txt = format_local(iso_utc)            # 'JJ/MM/AAAA HH:MM'
    return txt.split(" ")[-1] if txt else ""


def _plage(creneau: dict) -> str:
    """Libellé 'HH:MM–HH:MM' d'un créneau."""
    return f"{_heure(creneau['debut'])}–{_heure(creneau['fin'])}"


def _noms(affectations: list[dict]) -> str:
    """Concatène les noms affectés à une case (séparés par des retours ligne)."""
    return "\n".join(a["nom"] for a in affectations)


def construire_xlsx(grille: dict, nom_evenement: str) -> bytes:
    """
    Construit un classeur Excel : une feuille par jour (grille créneaux × postes)
    plus une feuille « Tâches » si des tâches ponctuelles existent.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    gras = Font(bold=True)
    grise = PatternFill("solid", fgColor="D9D9D9")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    haut = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    premiere = True
    postes = grille["postes"]

    for jour in grille["jours"]:
        ws = wb.active if premiere else wb.create_sheet()
        ws.title = (jour["libelle"] or "Jour")[:31]
        premiere = False

        ws["A1"] = f"{nom_evenement} — {jour['libelle']}"
        ws["A1"].font = gras
        # En-tête de colonnes : Horaire + un poste par colonne.
        ws.cell(row=2, column=1, value="Horaire").font = gras
        for j, p in enumerate(postes, start=2):
            c = ws.cell(row=2, column=j, value=p["nom"])
            c.font = gras
            c.alignment = centre

        for i, ligne in enumerate(jour["creneaux"], start=3):
            ws.cell(row=i, column=1, value=_plage(ligne["creneau"])).font = gras
            for j, case in enumerate(ligne["cases"], start=2):
                cell = ws.cell(row=i, column=j)
                if case["nb_requis"] <= 0:
                    cell.fill = grise            # case « grisée » (pas de besoin)
                else:
                    cell.value = _noms(case["affectations"])
                    cell.alignment = haut

        ws.column_dimensions["A"].width = 14
        for j in range(2, len(postes) + 2):
            ws.column_dimensions[ws.cell(row=2, column=j).column_letter].width = 18

    # Feuille des tâches ponctuelles (installation/rangement), si présentes.
    taches = [t for t in grille["taches"]]
    if taches:
        wt = wb.active if premiere else wb.create_sheet()
        wt.title = "Tâches"
        wt["A1"] = "Tâches ponctuelles"
        wt["A1"].font = gras
        wt.cell(row=2, column=1, value="Tâche").font = gras
        wt.cell(row=2, column=2, value="Horaire").font = gras
        wt.cell(row=2, column=3, value="Bénévoles").font = gras
        for i, t in enumerate(taches, start=3):
            c = t["creneau"]
            wt.cell(row=i, column=1, value=c.get("libelle") or "Tâche")
            wt.cell(row=i, column=2, value=f"{c['libelle_jour']} {_plage(c)}")
            wt.cell(row=i, column=3, value=", ".join(a["nom"] for a in t["affectations"]))
        for col, larg in (("A", 22), ("B", 22), ("C", 50)):
            wt.column_dimensions[col].width = larg

    if premiere:                                  # aucun jour : feuille vide lisible
        ws = wb.active
        ws.title = "Planning"
        ws["A1"] = f"{nom_evenement} — aucun créneau"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construire_pdf(grille: dict, nom_evenement: str) -> bytes:
    """
    Construit un PDF A4 paysage : un tableau par jour (créneaux × postes), plus
    un tableau des tâches ponctuelles. Pensé pour l'impression / l'affichage.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    styles = getSampleStyleSheet()
    petit = styles["BodyText"]
    petit.fontSize = 7
    petit.leading = 8

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4), title=f"Planning — {nom_evenement}",
        topMargin=1 * cm, bottomMargin=1 * cm, leftMargin=1 * cm, rightMargin=1 * cm,
    )
    elements = [Paragraph(f"Planning bénévole — {nom_evenement}", styles["Title"]),
                Spacer(1, 0.3 * cm)]
    postes = grille["postes"]
    largeur_totale = 27.0  # cm utiles en paysage
    larg_horaire = 3.0
    larg_poste = (largeur_totale - larg_horaire) / max(1, len(postes))

    def style_tableau(nb_lignes):
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a148c")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f0fa")]),
        ])

    for jour in grille["jours"]:
        elements.append(Paragraph(jour["libelle"], styles["Heading2"]))
        entetes = ["Horaire"] + [Paragraph(p["nom"], petit) for p in postes]
        lignes = [entetes]
        for ligne in jour["creneaux"]:
            row = [Paragraph(_plage(ligne["creneau"]), petit)]
            for case in ligne["cases"]:
                if case["nb_requis"] <= 0:
                    row.append("")              # case grisée
                else:
                    row.append(Paragraph(
                        "<br/>".join(a["nom"] for a in case["affectations"]) or "—",
                        petit))
            lignes.append(row)
        t = Table(lignes, colWidths=[larg_horaire * cm] + [larg_poste * cm] * len(postes),
                  repeatRows=1)
        t.setStyle(style_tableau(len(lignes)))
        # Grise les cellules sans besoin.
        for i, ligne in enumerate(jour["creneaux"], start=1):
            for j, case in enumerate(ligne["cases"], start=1):
                if case["nb_requis"] <= 0:
                    t.setStyle(TableStyle([("BACKGROUND", (j, i), (j, i),
                                            colors.HexColor("#d9d9d9"))]))
        elements.append(t)
        elements.append(Spacer(1, 0.4 * cm))

    if grille["taches"]:
        elements.append(Paragraph("Tâches ponctuelles", styles["Heading2"]))
        lignes = [["Tâche", "Horaire", "Bénévoles"]]
        for t in grille["taches"]:
            c = t["creneau"]
            lignes.append([
                Paragraph(c.get("libelle") or "Tâche", petit),
                Paragraph(f"{c['libelle_jour']} {_plage(c)}", petit),
                Paragraph(", ".join(a["nom"] for a in t["affectations"]) or "—", petit),
            ])
        tt = Table(lignes, colWidths=[6 * cm, 6 * cm, 15 * cm], repeatRows=1)
        tt.setStyle(style_tableau(len(lignes)))
        elements.append(tt)

    if not grille["jours"] and not grille["taches"]:
        elements.append(Paragraph("Aucun créneau défini.", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()
