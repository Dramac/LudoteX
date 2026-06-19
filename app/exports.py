"""
Génération des exports de statistiques en Excel (.xlsx) et PDF.

Les deux exports reçoivent le même dict `data` produit par
`services.collecter_stats` (synthèse + palmarès + liste détaillée des prêts), et
respectent donc le filtre de période actif. Ils renvoient des octets prêts à
être téléchargés (voir routes/stats.py).

Dépendances : openpyxl (Excel) et reportlab (PDF, déjà utilisé pour les planches
d'étiquettes).
"""

from __future__ import annotations

from io import BytesIO


def _libelle_metrique(metrique: str) -> str:
    return "par exemplaire" if metrique == "exemplaire" else "par total"


def construire_xlsx(data: dict, periode_txt: str) -> bytes:
    """
    Construit un classeur Excel à trois feuilles : Synthèse, Palmarès, Détail.

    Args:
        data: dict de services.collecter_stats.
        periode_txt: libellé lisible de la période (ou « toutes périodes »).

    Returns:
        Le contenu binaire du fichier .xlsx.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    gras = Font(bold=True)
    wb = Workbook()

    # --- Feuille 1 : Synthèse ---
    ws = wb.active
    ws.title = "Synthèse"
    g = data["globales"]
    ws["A1"] = "Statistiques de prêt — Des jeux plein la Manche"
    ws["A1"].font = gras
    ws["A2"] = f"Période : {periode_txt}"
    lignes = [
        ("Prêts au total", g["total_prets"]),
        ("Prêts en cours", g["en_cours"]),
        ("Titres prêtés", g["titres_pretes"]),
        ("Titres au catalogue", g["nb_titres"]),
        ("Durée moyenne de prêt", g.get("duree_moyenne", "—")),
    ]
    for i, (lib, val) in enumerate(lignes, start=4):
        ws[f"A{i}"] = lib
        ws[f"B{i}"] = val

    # --- Feuille 2 : Palmarès (plus puis moins prêtés) ---
    wp = wb.create_sheet("Palmarès")
    wp["A1"] = f"Palmarès ({_libelle_metrique(data['metrique'])})"
    wp["A1"].font = gras
    ligne = 3
    for titre_section, jeux in (("Les plus prêtés", data["plus"]),
                                ("Les moins prêtés", data["moins"])):
        wp[f"A{ligne}"] = titre_section
        wp[f"A{ligne}"].font = gras
        ligne += 1
        for col, entete in enumerate(["Jeu", "Prêts", "Exemplaires", "Par exempl."]):
            cell = wp.cell(row=ligne, column=1 + col, value=entete)
            cell.font = gras
        ligne += 1
        for jeu in jeux:
            wp.cell(row=ligne, column=1, value=jeu["nom"])
            wp.cell(row=ligne, column=2, value=jeu["nb_prets"])
            wp.cell(row=ligne, column=3, value=jeu["nb_exemplaires"])
            wp.cell(row=ligne, column=4, value=round(jeu["par_exemplaire"], 2))
            ligne += 1
        ligne += 1

    # --- Feuille 3 : Détail des prêts ---
    wd = wb.create_sheet("Détail")
    for col, entete in enumerate(["Jeu", "Exemplaire", "Sortie", "Retour",
                                  "Durée", "N° emplacement"]):
        c = wd.cell(row=1, column=1 + col, value=entete)
        c.font = gras
    for i, p in enumerate(data["prets"], start=2):
        wd.cell(row=i, column=1, value=p["nom"])
        wd.cell(row=i, column=2, value=p["id_exemplaire"])
        wd.cell(row=i, column=3, value=p["sortie_locale"])
        wd.cell(row=i, column=4, value=p["retour_local"] or "en cours")
        wd.cell(row=i, column=5, value=p["duree_txt"])
        wd.cell(row=i, column=6, value=p["numero_pochette"])

    # Largeurs de colonnes lisibles.
    for feuille, largeurs in ((ws, [24, 14]), (wp, [40, 10, 12, 12]),
                              (wd, [40, 12, 18, 18, 12, 14])):
        for idx, larg in enumerate(largeurs):
            feuille.column_dimensions[chr(65 + idx)].width = larg

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Sections possibles du PDF, dans l'ordre d'apparition.
SECTIONS_PDF = ("synthese", "plus", "moins", "detail")


def construire_pdf(data: dict, periode_txt: str,
                   sections: "set[str] | None" = None) -> bytes:
    """
    Construit un PDF de bilan, avec sections au choix.

    Args:
        data: dict de services.collecter_stats.
        periode_txt: libellé lisible de la période.
        sections: ensemble des sections à inclure parmi SECTIONS_PDF
            ("synthese", "plus", "moins", "detail"). None = toutes.

    Returns:
        Le contenu binaire du fichier .pdf.
    """
    if sections is None:
        sections = set(SECTIONS_PDF)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="Statistiques de prêt",
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elements = []

    def tableau(entetes, lignes, largeurs):
        t = Table([entetes] + lignes, colWidths=largeurs, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a148c")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f0fa")]),
        ]))
        return t

    elements.append(Paragraph("Statistiques de prêt — Des jeux plein la Manche",
                              styles["Title"]))
    elements.append(Paragraph(f"Période : {periode_txt}", styles["Normal"]))
    elements.append(Spacer(1, 0.4 * cm))

    # Synthèse
    if "synthese" in sections:
        g = data["globales"]
        elements.append(Paragraph("Synthèse", styles["Heading2"]))
        elements.append(tableau(
            ["Indicateur", "Valeur"],
            [["Prêts au total", str(g["total_prets"])],
             ["Prêts en cours", str(g["en_cours"])],
             ["Titres prêtés", str(g["titres_pretes"])],
             ["Titres au catalogue", str(g["nb_titres"])],
             ["Durée moyenne de prêt", g.get("duree_moyenne", "—")]],
            [8 * cm, 4 * cm]))
        elements.append(Spacer(1, 0.4 * cm))

    # Palmarès (les deux sections sont indépendantes)
    palmares_sections = []
    if "plus" in sections:
        palmares_sections.append(("Les plus prêtés", data["plus"]))
    if "moins" in sections:
        palmares_sections.append(("Les moins prêtés", data["moins"]))
    for titre_section, jeux in palmares_sections:
        elements.append(Paragraph(
            f"{titre_section} ({_libelle_metrique(data['metrique'])})",
            styles["Heading2"]))
        lignes = [[j["nom"], str(j["nb_prets"]), str(j["nb_exemplaires"]),
                   f"{j['par_exemplaire']:.2f}"] for j in jeux]
        elements.append(tableau(["Jeu", "Prêts", "Ex.", "Par ex."],
                                lignes, [9 * cm, 2 * cm, 2 * cm, 2.5 * cm]))
        elements.append(Spacer(1, 0.4 * cm))

    # Détail des prêts
    if "detail" in sections:
        elements.append(Paragraph(f"Détail des prêts ({len(data['prets'])})",
                                  styles["Heading2"]))
        lignes = [[p["nom"], p["id_exemplaire"], p["sortie_locale"],
                   p["retour_local"] or "en cours", p["duree_txt"],
                   str(p["numero_pochette"])]
                  for p in data["prets"]]
        if lignes:
            elements.append(tableau(
                ["Jeu", "Ex.", "Sortie", "Retour", "Durée", "Empl."],
                lignes,
                [5.5 * cm, 1.8 * cm, 3 * cm, 3 * cm, 2 * cm, 1.4 * cm]))
        else:
            elements.append(Paragraph("Aucun prêt sur la période.", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()
